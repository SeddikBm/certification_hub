#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script d'importation des 53 certifications, squads et liens Udemy depuis le fichier Excel :
Plan_Formation_IT_2025_2027_PRODUCTION_v1.0.xlsm

Ce script applique strictement les règles de normalisation :
- Colonnes dédiées : code, name, provider, difficulty, priority, exam_cost_usd, validity_months, official_url, exam_provider_url
- Metadata JSONB : uniquement les valeurs sans colonne dédiée (price_mad, preparation_hours, business_value, version, squad_domain, squads_affected, etc.)
- Squads : Création/Upsert des 5 squads avec leurs color_hex
- certification_squads : Liaisons complètes avec priorité (1=Obligatoire, 3=Recommandé, 5=Optionnel)
"""

import sys
import os
import json
import re
import openpyxl
import subprocess
from pathlib import Path

# Chemins
BASE_DIR = Path(__file__).resolve().parent.parent
EXCEL_FILE = BASE_DIR / "Plan_Formation_IT_2025_2027_PRODUCTION_v1.0.xlsm"
SQL_OUTPUT_FILE = BASE_DIR / "scripts" / "import_catalog.sql"
MIGRATION_FILE = BASE_DIR / "backend-spring" / "src" / "main" / "resources" / "db" / "migration" / "V8__import_catalog_53_certifications.sql"

# 1. Définition des 5 Squads officielles
SQUADS = [
    {
        "id": "a0000000-0000-0000-0000-000000000001",
        "name": ".NET Squad",
        "aliases": [".net", ".net squad", "dotnet"],
        "description": "Équipe .NET",
        "color_hex": "#5C2D91"
    },
    {
        "id": "a0000000-0000-0000-0000-000000000002",
        "name": "Java Squad",
        "aliases": ["java", "java squad"],
        "description": "Équipe Java",
        "color_hex": "#E76F00"
    },
    {
        "id": "a0000000-0000-0000-0000-000000000003",
        "name": "DevOps Squad",
        "aliases": ["devops", "devops squad"],
        "description": "Équipe DevOps",
        "color_hex": "#0078D4"
    },
    {
        "id": "a0000000-0000-0000-0000-000000000004",
        "name": "Data Squad",
        "aliases": ["data", "data squad", "data / ai", "data/ai"],
        "description": "Équipe Data / AI",
        "color_hex": "#00A36C"
    },
    {
        "id": "a0000000-0000-0000-0000-000000000005",
        "name": "QA-Testing",
        "aliases": ["qa-testing", "qa", "testing", "qualite", "qualité"],
        "description": "Équipe Qualité & Test",
        "color_hex": "#D13438"
    }
]

# 2. Providers standards reconnus
PROVIDERS_MAP = {
    'PSM-I': 'Scrum.org',
    'PSPO-I': 'Scrum.org',
    'SA': 'Scaled Agile',
    'RTE': 'Scaled Agile',
    'PAL-I': 'Scrum.org',
    'ITIL4-F': 'PeopleCert / Axelos',
    'OCP-21': 'Oracle',
    'SPRING-PRO': 'VMware / Broadcom',
    'AZ-900': 'Microsoft',
    'AZ-204': 'Microsoft',
    'AZ-104': 'Microsoft',
    'AZ-305': 'Microsoft',
    'AZ-400': 'Microsoft',
    'VCP-DCV': 'VMware',
    'RHCSA': 'Red Hat',
    'RHCE': 'Red Hat',
    'LFCS': 'Linux Foundation',
    'GITLAB-PROF': 'GitLab',
    'CJE': 'CloudBees',
    'PCA': 'CNCF',
    'CCNA': 'Cisco',
    'CKAD': 'CNCF',
    'CKA': 'CNCF',
    'CKS': 'CNCF',
    'TERRAFORM-003': 'HashiCorp',
    'SAA-C03': 'AWS',
    'SAP-C02': 'AWS',
    'DVA-C02': 'AWS',
    'ACE': 'Google Cloud',
    'TOGAF-10': 'The Open Group',
    'AI-900': 'Microsoft',
    'AI-102': 'Microsoft',
    'SY0-701': 'CompTIA',
    'CEH': 'EC-Council',
    'PMP': 'PMI',
    'PRINCE2-P': 'PeopleCert / Axelos',
    'CSM': 'Scrum Alliance',
    'DCEA': 'Databricks',
    'CCDAK': 'Confluent',
    'CCAAK': 'Confluent',
    'CDP-0011': 'Cloudera',
    'CAOP': 'Cloudera',
    'CDE': 'Cloudera',
    'DAS-C01': 'AWS',
    'DP-203': 'Microsoft',
    'CTFL': 'ISTQB',
    'CT-TAE': 'ISTQB',
    'SELENIUM': 'Selenium',
    'CT-AT': 'ISTQB',
    'CYPRESS': 'Cypress',
    'POSTMAN': 'Postman',
    'JMETER': 'Apache',
    'CTAL-TM': 'ISTQB'
}

# 3. Mappings de normalisation
DIFFICULTY_MAP = {
    'Junior': 'FOUNDATIONAL',
    'Senior': 'INTERMEDIATE',
    'Lead': 'ADVANCED',
    'Expert': 'EXPERT',
    'Manager': 'ADVANCED'
}

PRIORITY_MAP = {
    'OBLIGATOIRE': 'MANDATORY',
    'RECOMMANDE': 'HIGH',
    'OPTIONNEL': 'NORMAL'
}

VALIDITY_MONTHS_MAP = {
    '1 an': 12,
    '2 ans': 24,
    '3 ans': 36,
    '5 ans': 60,
    'Permanente': None
}

SQUAD_PRIORITY_MAP = {
    'OBLIGATOIRE': 1,
    'RECOMMANDE': 3,
    'OPTIONNEL': 5
}

def escape_sql_str(val):
    if val is None:
        return 'NULL'
    return "'" + str(val).replace("'", "''") + "'"

def parse_excel():
    if not EXCEL_FILE.exists():
        print(f"Erreur : Fichier introuvable à {EXCEL_FILE}")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(EXCEL_FILE), data_only=False)
    if 'Catalogue Enrichi' not in wb.sheetnames:
        print("Erreur : Feuille 'Catalogue Enrichi' introuvable")
        sys.exit(1)

    ws = wb['Catalogue Enrichi']
    certifications = []

    # Parcours des lignes de certification (lignes 4 à 56)
    for r in range(4, ws.max_row + 1):
        domain = ws.cell(r, 1).value
        if domain in ['INSTRUCTIONS:', 'LEGENDE'] or not domain:
            continue

        squads_affected_raw = ws.cell(r, 2).value
        cert_name = ws.cell(r, 3).value
        code = ws.cell(r, 4).value

        if not cert_name or not code:
            continue

        version = ws.cell(r, 5).value
        price_usd = ws.cell(r, 6).value
        price_mad = ws.cell(r, 7).value
        
        # Hyperliens
        cell_udemy = ws.cell(r, 8)
        cell_official = ws.cell(r, 9)
        udemy_url = cell_udemy.hyperlink.target if cell_udemy.hyperlink else None
        official_url = cell_official.hyperlink.target if cell_official.hyperlink else None

        validity_str = str(ws.cell(r, 10).value or '').strip()
        category = str(ws.cell(r, 11).value or '').strip()
        hours = ws.cell(r, 12).value
        business_value = str(ws.cell(r, 13).value or '').strip()
        level = str(ws.cell(r, 14).value or '').strip()

        # Normalisation
        provider = PROVIDERS_MAP.get(code)
        if not provider:
            # Fallback deduction
            name_upper = cert_name.upper()
            if 'AWS' in name_upper or 'AMAZON' in name_upper: provider = 'AWS'
            elif 'AZURE' in name_upper or 'MICROSOFT' in name_upper: provider = 'Microsoft'
            elif 'GOOGLE' in name_upper or 'GCP' in name_upper: provider = 'Google Cloud'
            elif 'ISTQB' in name_upper: provider = 'ISTQB'
            elif 'SCRUM' in name_upper: provider = 'Scrum.org'
            elif 'RED HAT' in name_upper: provider = 'Red Hat'
            else: provider = domain

        difficulty = DIFFICULTY_MAP.get(level, 'INTERMEDIATE')
        priority = PRIORITY_MAP.get(category, 'NORMAL')
        validity_months = VALIDITY_MONTHS_MAP.get(validity_str)
        if validity_months is None and 'an' in validity_str:
            # ex: '2 ans'
            match = re.search(r'(\d+)', validity_str)
            if match:
                validity_months = int(match.group(1)) * 12

        # Nettoyage des coûts
        try:
            exam_cost_usd = float(price_usd) if price_usd is not None else 0.0
        except (ValueError, TypeError):
            exam_cost_usd = 0.0

        try:
            cost_mad = float(price_mad) if price_mad is not None else 0.0
        except (ValueError, TypeError):
            cost_mad = 0.0

        try:
            prep_hours = int(hours) if hours is not None else None
        except (ValueError, TypeError):
            prep_hours = None

        # Nettoyage Nom de certification (gestion des caractères spéciaux)
        clean_name = cert_name.replace('\ufffd', '–').strip()

        # Construction Metadata JSONB : uniquement les valeurs sans colonne dédiée
        metadata = {
            "price_mad": cost_mad,
            "preparation_hours": prep_hours,
            "business_value": business_value,
            "level": level,
            "category": category,
            "version": str(version) if version is not None else None,
            "squad_domain": domain,
            "squads_affected": squads_affected_raw
        }

        # Détermination des squads associées
        associated_squad_ids = []
        squad_prio = SQUAD_PRIORITY_MAP.get(category, 3)

        aff_lower = str(squads_affected_raw or '').lower().strip()
        if aff_lower == 'tous':
            associated_squad_ids = [sq['id'] for sq in SQUADS]
        else:
            parts = [p.strip() for p in aff_lower.split(',') if p.strip()]
            for part in parts:
                matched = False
                for sq in SQUADS:
                    if part in sq['aliases'] or sq['name'].lower() == part:
                        if sq['id'] not in associated_squad_ids:
                            associated_squad_ids.append(sq['id'])
                        matched = True
                        break
                if not matched:
                    # Partial match
                    for sq in SQUADS:
                        if sq['aliases'][0] in part:
                            if sq['id'] not in associated_squad_ids:
                                associated_squad_ids.append(sq['id'])
                            break

        certifications.append({
            "code": code,
            "name": clean_name,
            "provider": provider,
            "difficulty": difficulty,
            "priority": priority,
            "exam_cost_usd": exam_cost_usd,
            "validity_months": validity_months,
            "official_url": udemy_url,
            "exam_provider_url": official_url,
            "metadata": metadata,
            "squad_ids": associated_squad_ids,
            "squad_priority": squad_prio
        })

    return certifications

def generate_sql(certifications):
    lines = []
    lines.append("-- ==============================================================================")
    lines.append("-- IMPORT CATALOGUE : 53 CERTIFICATIONS, 5 SQUADS & LIAISONS UDEMY/OFFICIELLES")
    lines.append("-- Généré automatiquement depuis Plan_Formation_IT_2025_2027_PRODUCTION_v1.0.xlsm")
    lines.append("-- ==============================================================================\n")

    lines.append("BEGIN;\n")

    # 1. Upsert des 5 Squads
    lines.append("-- 1. Création / Mise à jour des 5 Squads")
    for sq in SQUADS:
        lines.append(f"""INSERT INTO squads (id, name, description, color_hex)
VALUES ('{sq['id']}', {escape_sql_str(sq['name'])}, {escape_sql_str(sq['description'])}, {escape_sql_str(sq['color_hex'])})
ON CONFLICT (id) DO UPDATE SET
    name = EXCLUDED.name,
    description = EXCLUDED.description,
    color_hex = EXCLUDED.color_hex,
    updated_at = NOW();""")
    lines.append("")

    # 2. Insert/Upsert des 53 Certifications
    lines.append("-- 2. Insertion / Mise à jour des 53 Certifications")
    for cert in certifications:
        validity_val = str(cert['validity_months']) if cert['validity_months'] is not None else "NULL"
        metadata_json = json.dumps(cert['metadata'], ensure_ascii=False)
        
        lines.append(f"""INSERT INTO certifications (
    code, name, provider, difficulty, priority,
    exam_cost_usd, training_cost_usd, validity_months,
    official_url, exam_provider_url, metadata
)
VALUES (
    {escape_sql_str(cert['code'])},
    {escape_sql_str(cert['name'])},
    {escape_sql_str(cert['provider'])},
    '{cert['difficulty']}'::cert_difficulty,
    '{cert['priority']}'::cert_priority,
    {cert['exam_cost_usd']:.2f},
    0.00,
    {validity_val},
    {escape_sql_str(cert['official_url'])},
    {escape_sql_str(cert['exam_provider_url'])},
    {escape_sql_str(metadata_json)}::jsonb
)
ON CONFLICT (code) DO UPDATE SET
    name = EXCLUDED.name,
    provider = EXCLUDED.provider,
    difficulty = EXCLUDED.difficulty,
    priority = EXCLUDED.priority,
    exam_cost_usd = EXCLUDED.exam_cost_usd,
    validity_months = EXCLUDED.validity_months,
    official_url = EXCLUDED.official_url,
    exam_provider_url = EXCLUDED.exam_provider_url,
    metadata = EXCLUDED.metadata,
    deleted_at = NULL,
    updated_at = NOW();""")
    lines.append("")

    # 3. Liaisons certification_squads
    lines.append("-- 3. Liaisons certification_squads")
    for cert in certifications:
        code_esc = escape_sql_str(cert['code'])
        for squad_id in cert['squad_ids']:
            lines.append(f"""INSERT INTO certification_squads (certification_id, squad_id, priority)
SELECT c.id, '{squad_id}'::uuid, {cert['squad_priority']}
FROM certifications c
WHERE c.code = {code_esc}
ON CONFLICT (certification_id, squad_id) DO UPDATE SET
    priority = EXCLUDED.priority;""")

    lines.append("\nCOMMIT;\n")
    return "\n".join(lines)

def main():
    print("Lecture du fichier Excel...")
    certs = parse_excel()
    print(f"{len(certs)} certifications extraites avec succès.")

    sql_content = generate_sql(certs)

    # Écriture du script SQL
    SQL_OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(SQL_OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(sql_content)
    print(f"Fichier SQL généré : {SQL_OUTPUT_FILE}")

    # Écriture de la migration Flyway V8
    MIGRATION_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(MIGRATION_FILE, "w", encoding="utf-8") as f:
        f.write(sql_content)
    print(f"Migration Flyway générée : {MIGRATION_FILE}")

    # Exécution optionnelle ou directe si demandé
    if "--execute" in sys.argv:
        print("\nExécution du SQL dans la base de données Docker certifhub_db...")
        cmd = ["docker", "exec", "-i", "certifhub_db", "psql", "-U", "certif_user", "-d", "certifhub_db"]
        res = subprocess.run(cmd, input=sql_content.encode("utf-8"), capture_output=True)
        if res.returncode == 0:
            print("Importation exécutée avec succès dans PostgreSQL !")
            print(res.stdout.decode("utf-8", errors="ignore"))
        else:
            print("Erreur lors de l'exécution SQL :", res.stderr.decode("utf-8", errors="ignore"))
            sys.exit(1)

if __name__ == "__main__":
    main()
